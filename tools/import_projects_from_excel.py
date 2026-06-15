import argparse
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

import openpyxl


PROJECT_COLUMNS = [
    "tracking_id",
    "Created_Date",
    "khach_hang",
    "nhan_vien_kinh_doanh",
    "ten_san_pham",
    "quy_cach",
    "nguoi_lien_he_kh",
    "so_luong",
    "ma_po",
    "ma_ban_ve",
    "ma_ban_ve_ky_thuat",
    "ma_me",
    "loai_san_pham",
    "nhan_vien_thiet_ke",
    "tinh_trang_hoan_thanh",
    "urgency_level",
    "thoi_gian_mong_muon_ban_ve",
    "thoi_gian_hoan_thanh_ke_hoach",
    "sales_name",
    "user_id",
    "is_pending",
    "accepted_by",
    "accepted_at",
    "desired_solution_time",
]


def clean_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        return value.strip()
    return value


def normalize_urgency(value):
    text = str(clean_value(value))
    lowered = text.lower()
    if not text:
        return ""
    if "非常" in text or "very" in lowered or "rất" in lowered:
        return "very_urgent"
    if "紧急" in text or "urgent" in lowered or "khẩn" in lowered:
        return "urgent"
    if "正常" in text or "normal" in lowered or "bình" in lowered:
        return "normal"
    return text


def is_project_row(values):
    # Ignore rows that only carry the month marker in column B.
    return any(clean_value(values[index]) not in ("", " ") for index in range(2, 19))


def read_projects(excel_path):
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    worksheet = workbook.worksheets[0]
    projects = []

    for row_index in range(3, worksheet.max_row + 1):
        values = [worksheet.cell(row_index, col_index).value for col_index in range(1, 20)]
        if not is_project_row(values):
            continue

        project = {
            "tracking_id": len(projects) + 1,
            "Created_Date": clean_value(values[1]),
            "khach_hang": clean_value(values[2]),
            "nhan_vien_kinh_doanh": clean_value(values[3]),
            "ten_san_pham": clean_value(values[4]),
            "quy_cach": clean_value(values[5]),
            "nguoi_lien_he_kh": clean_value(values[6]),
            "so_luong": clean_value(values[7]),
            "ma_po": clean_value(values[8]),
            "ma_ban_ve": clean_value(values[9]),
            "ma_ban_ve_ky_thuat": clean_value(values[10]),
            "ma_me": clean_value(values[11]),
            "loai_san_pham": clean_value(values[12]),
            "nhan_vien_thiet_ke": clean_value(values[13]),
            "tinh_trang_hoan_thanh": clean_value(values[14]),
            "urgency_level": normalize_urgency(values[15]),
            "thoi_gian_mong_muon_ban_ve": clean_value(values[17]),
            "thoi_gian_hoan_thanh_ke_hoach": clean_value(values[18]),
            "sales_name": clean_value(values[3]),
            "user_id": None,
            "is_pending": "no",
            "accepted_by": "",
            "accepted_at": clean_value(values[16]),
            "desired_solution_time": "",
        }
        projects.append(project)

    return worksheet.title, projects


def ensure_customers_table(cursor):
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code VARCHAR(50),
            name VARCHAR(200) UNIQUE NOT NULL,
            phonetic VARCHAR(100),
            english_name VARCHAR(200),
            contact_person VARCHAR(100),
            phone VARCHAR(50),
            email VARCHAR(100),
            address TEXT
        )
        """
    )


def import_projects(db_path, projects):
    connection = sqlite3.connect(db_path)
    cursor = connection.cursor()
    cursor.execute("BEGIN")

    cursor.execute("DELETE FROM projects")
    placeholders = ", ".join(["?"] * len(PROJECT_COLUMNS))
    columns_sql = ", ".join(PROJECT_COLUMNS)

    for project in projects:
        cursor.execute(
            f"INSERT INTO projects ({columns_sql}) VALUES ({placeholders})",
            [project.get(column) for column in PROJECT_COLUMNS],
        )

    ensure_customers_table(cursor)
    customer_names = sorted({project["khach_hang"] for project in projects if project["khach_hang"]})
    for customer_name in customer_names:
        cursor.execute("INSERT OR IGNORE INTO customers (name) VALUES (?)", (customer_name,))

    connection.commit()
    connection.close()
    return len(customer_names)


def main():
    parser = argparse.ArgumentParser(description="Import the first sheet of a projects Excel file into DB.db.")
    parser.add_argument("excel_path", type=Path)
    parser.add_argument("--db", type=Path, default=Path("DB.db"))
    parser.add_argument("--no-backup", action="store_true")
    args = parser.parse_args()

    excel_path = args.excel_path.resolve()
    db_path = args.db.resolve()
    if not excel_path.exists():
        raise FileNotFoundError(excel_path)
    if not db_path.exists():
        raise FileNotFoundError(db_path)

    if not args.no_backup:
        backup_path = db_path.with_name(f"{db_path.stem}.backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}{db_path.suffix}")
        shutil.copy2(db_path, backup_path)
        print(f"Backup: {backup_path}")

    sheet_name, projects = read_projects(excel_path)
    customer_count = import_projects(db_path, projects)
    print(f"Sheet: {sheet_name}")
    print(f"Imported projects: {len(projects)}")
    print(f"Upserted customers: {customer_count}")


if __name__ == "__main__":
    main()
